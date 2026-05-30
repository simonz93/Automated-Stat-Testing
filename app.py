"""
Significance Triangle Sync — Streamlit web app
==============================================
Upload a survey data Excel file and a PowerPoint deck; the app inserts ▲/▼
significance markers into the deck's charts, tables, and text boxes, then lets
you download the updated deck plus a progress report.

Flow:
  1. Configure the Excel layout (sheet + where the header/letter/label rows are).
  2. The app scans the hierarchical column headers and offers cascading dropdowns
     for the default column cut.
  3. Optionally add per-shape overrides (same cascading picker per row).
  4. Upload the PowerPoint and run.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py
"""

import io
import inspect
import openpyxl
import pandas as pd
import streamlit as st

import triangle_engine as engine


st.set_page_config(page_title="Significance Triangle Sync", page_icon="▲", layout="wide")

# ── Engine compatibility guard ────────────────────────────────────────────────
# app.py and triangle_engine.py are a matched pair. If an OLD engine is imported
# (e.g. a stale copy elsewhere on the path) the run would fail deep inside with a
# confusing error. Detect that up front and say exactly which file was loaded.
_engine_path = getattr(engine, "__file__", "(unknown)")
_run_sync_params = inspect.signature(engine.run_sync).parameters
if "worksheet" not in _run_sync_params:
    st.error(
        "The triangle_engine.py being loaded is an OLD version that doesn't match "
        "this app.py.\n\n"
        f"Python loaded the engine from:\n`{_engine_path}`\n\n"
        "Replace that file with the matching triangle_engine.py (the one delivered "
        "alongside this app.py), delete any `__pycache__` folder next to it, and "
        "restart. Both files must come from the same release."
    )
    st.stop()

st.title("▲ Significance Triangle Sync")
st.caption(
    "Insert ▲/▼ significance markers into a PowerPoint deck from survey "
    "cross-tab data. Shapes are matched by their Selection-Pane names "
    "(those starting with “x ”)."
)

# ── Help buttons ──────────────────────────────────────────────────────────────
hb1, hb2, _ = st.columns([1.2, 1.4, 4])
with hb1:
    if st.button("📖 How to use", help="Step-by-step instructions"):
        st.session_state["show_help"] = not st.session_state.get("show_help", False)
with hb2:
    if st.button("🎛 PowerPoint set-up", help="Shape-naming syntax reference"):
        st.session_state["show_syntax"] = not st.session_state.get("show_syntax", False)

if st.session_state.get("show_help"):
    with st.expander("How to use this app", expanded=True):
        st.markdown(
            """
**What it does** — reads a survey cross-tab Excel, compares Wave 4 vs Wave 5 for
the column cut you choose, and writes ▲ (significant increase) / ▼ (significant
decrease) markers into the matching shapes of your PowerPoint deck.

**Steps**

1. **Upload the data Excel and the PowerPoint** at the top. To swap a file,
   upload a different one in the same box.
2. **Step 1 · Excel layout** — pick the data sheet and confirm where the header
   rows, letter row, and label column live, then press **Continue → scan this
   sheet**. Only the sheet you choose is scanned, which keeps it fast. If you
   change any layout field, press Continue again to re-scan.
3. **Step 2 · Default column cut** — choose the column the tool reads for every
   shape. Each level unlocks the next (e.g. Country → Rest of the world → — →
   Total). Set the minimum % point difference and any slides to skip here too.
4. **Step 3 · Shape overrides (optional)** — for shapes that should read a
   *different* column than the default, add a row: slide number, the exact shape
   name, and the column path.
5. **Step 4 · Run** — press **Apply triangles**. When it finishes you can
   download the updated PowerPoint and an Excel report. Downloading no longer
   clears the screen; press **↻ Run again** to apply again with the current
   settings.

**Reading the report** — each shape is one of:
- 🟩 **Data found – stat testing applied** — a ▲/▼ was written.
- ⬜ **Data found – not significant** — matched the data, but the Wave 4↔5 gap
  was below your threshold, so no marker.
- 🟥 **Data not found** — the shape's label/number didn't match the chosen
  column, so nothing was written (check the shape name or column cut).

**Validation** — a marker is only applied when the number shown in the shape
matches the chosen column's Wave 5 value (charts are checked to 2 decimals).
This prevents markers landing on shapes that are actually showing a different
column.
            """
        )

if st.session_state.get("show_syntax"):
    with st.expander("PowerPoint set-up — shape-naming syntax", expanded=True):
        st.markdown(
            """
Name shapes in the **Selection Pane** (Home → Arrange → Selection Pane). Every
shape the tool should touch starts with **`x `** (lower-case x + space) followed
by a **section code** and a dot. The code is the question/section label in
column 1 of the Excel (e.g. `B2`, `D6_1`, `C3_1`).
            """
        )
        st.markdown("**1 · Single-code shapes (charts, tables, simple text)**")
        st.code(
            "x B2.\n"
            "x D6.\n"
            "x B2. (2)        ← a trailing \"(2)\" duplicate marker is ignored",
            language=None,
        )
        st.caption(
            "The tool matches each chart category / table row / series to a data "
            "row in that section by its label, and applies the marker where the "
            "shown number matches the chosen column."
        )

        st.markdown("**2 · Exact-label text boxes**")
        st.code(
            "x B1. Net: Very/somewhat satisfied\n"
            "x D1_1. Net: Total Aware",
            language=None,
        )
        st.caption(
            "Everything after the code and dot is the exact row label to look up "
            "in that section. The marker is written into the text box."
        )

        st.markdown("**3 · Rank text boxes** — `x CODE_n. <rank>`")
        st.code(
            "x D6_1. 1     ← the #1 row when CODE's Wave 5 values are sorted high→low\n"
            "x D6_1. 2\n"
            "x D6_1. 3",
            language=None,
        )
        st.caption(
            "Use when a row of text boxes shows the top-N items by value. Rank 1 "
            "is the highest Wave 5 value in that section, rank 2 the next, etc."
        )

        st.markdown("**4 · LINK text boxes** — `x CODE_n. LINK OTHER. <rank>`")
        st.code(
            "x C3_1. LINK PRIMARY_MODE. 1\n"
            "x C3_1. LINK PRIMARY_MODE. 2\n"
            "x C3_1. LINK PRIMARY_MODE. 3",
            language=None,
        )
        st.caption(
            "Like a rank box, but the rank order follows ANOTHER section's sort. "
            "Rank N takes the label sitting at OTHER's rank-N position, then shows "
            "THIS section's value and marker for that same label. Example: the "
            "C3_1 boxes are laid out in PRIMARY_MODE's order rather than C3_1's own."
        )

        st.markdown("**5 · xMULTI charts** — `xMULTI ROOT.`")
        st.code(
            "xMULTI D6.\n"
            "xMULTI A2.",
            language=None,
        )
        st.caption(
            "For grouped/clustered charts that combine several sub-codes under one "
            "root (e.g. D6_1, D6_2 …). The tool reads each bar's section from the "
            "category or series prefix and matches automatically — both layouts "
            "(code-on-category and code-on-series) are handled."
        )

        st.markdown("**Tips**")
        st.markdown(
            "- Names are matched exactly, so keep the `x `, the code, and the dot.\n"
            "- A shape with no dot, or a code not present in the data, is reported "
            "as **Data not found** and left untouched.\n"
            "- To make one shape read a different column than the default, leave "
            "its name as-is and add it under **Step 3 · Shape overrides**."
        )



# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_int_list(raw):
    return [int(p.strip()) for p in str(raw).split(",") if p.strip().lstrip("-").isdigit()]


@st.cache_data(show_spinner=False)
def list_sheets(file_bytes):
    # read_only is lazy: listing names parses no cell data, so this is fast.
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, keep_links=False)
    names = list(wb.sheetnames)
    wb.close()
    return names


@st.cache_resource(show_spinner=False)
def load_worksheet(file_bytes_hash, _file_bytes, sheet):
    """
    Load the workbook ONCE and return the target worksheet (a live object,
    cached across reruns). `file_bytes_hash` is the cache key; `_file_bytes`
    is excluded from hashing (leading underscore) since it can be large.
    A regular (non-read_only) load is used because per-cell access afterwards
    is far faster than read_only mode for our scanning.
    """
    wb = openpyxl.load_workbook(io.BytesIO(_file_bytes), data_only=True, keep_links=False)
    return wb[sheet]


@st.cache_data(show_spinner=False)
def build_tree(file_bytes_hash, _ws_key, header_rows, letter_row, label_col,
               _ws):
    """Build the header tree from an already-loaded worksheet under the layout."""
    engine.HEADER_ROWS = list(header_rows)
    engine.LETTER_ROW  = int(letter_row)
    engine.LABEL_COL   = int(label_col)
    engine.get_col_ancestry.__defaults__[0].clear()
    return engine.build_header_tree(_ws)


def cascade_options(paths, chosen):
    """
    Given the list of full path tuples and the values chosen so far (a list,
    one per level, '' = not chosen yet), return the valid options for the NEXT
    unchosen level — i.e. the distinct values at that level among paths whose
    earlier levels match what's chosen.
    """
    level = len(chosen)
    opts = []
    for p in paths:
        if level >= len(p):
            continue
        if all(p[i] == chosen[i] for i in range(level)):
            opts.append(p[level])
    # Preserve order, drop dups
    seen, out = set(), []
    for o in opts:
        if o not in seen:
            seen.add(o); out.append(o)
    return out


def cascade_picker(paths, levels, key_prefix, defaults=None):
    """
    Render `levels` cascading selectboxes. Each becomes active only once the
    previous one is chosen. Returns the chosen path as a list (may contain ''
    for legitimately blank levels). Blank ('') options are shown as “—”.
    """
    BLANK = "—"
    chosen = []
    defaults = defaults or []
    for lvl in range(levels):
        opts = cascade_options(paths, chosen)
        if not opts:
            break
        display = [BLANK if o == "" else o for o in opts]
        # Default selection if provided and valid at this level
        idx = 0
        if lvl < len(defaults):
            want = defaults[lvl]
            disp_want = BLANK if want == "" else want
            if disp_want in display:
                idx = display.index(disp_want)
        sel = st.selectbox(
            f"Level {lvl + 1}", display, index=idx,
            key=f"{key_prefix}_lvl{lvl}",
        )
        chosen.append("" if sel == BLANK else sel)
    return chosen


# ── Top progress bar (item 2) ─────────────────────────────────────────────────
# A real animated bar fixed across the very top of the page. Rendered "active"
# (visible + animating) only while a heavy step is running this rerun.
_TOP_BAR_CSS = """
<style>
  #top-progress-bar {
    position: fixed; top: 0; left: 0; right: 0; height: 6px;
    z-index: 100000; background: rgba(31,78,121,0.10); overflow: hidden;
  }
  #top-progress-bar .fill {
    position: absolute; top: 0; height: 100%; width: 30%;
    background: linear-gradient(90deg,#1F4E79,#3A9262,#1F4E79);
    animation: tp-slide 1.1s ease-in-out infinite;
  }
  @keyframes tp-slide {
    0%   { left: -30%; width: 30%; }
    50%  { left: 40%;  width: 45%; }
    100% { left: 100%; width: 30%; }
  }
</style>
"""

def _top_bar(active: bool):
    """Render the fixed top progress bar; when active it animates."""
    if active:
        st.markdown(_TOP_BAR_CSS + '<div id="top-progress-bar"><div class="fill"></div></div>',
                    unsafe_allow_html=True)


# ── File uploaders (one each) with a refresh control (item 3) ─────────────────
left, right = st.columns(2)
with left:
    excel_file = st.file_uploader(
        "Data Excel (.xlsx)", type=["xlsx"], accept_multiple_files=False,
        key="excel_uploader")
with right:
    pptx_file = st.file_uploader(
        "PowerPoint (.pptx)", type=["pptx"], accept_multiple_files=False,
        key="pptx_uploader")

if excel_file is None:
    st.info("Upload the data Excel to begin — the column settings are built from it.")
    st.stop()

excel_bytes = excel_file.getvalue()
try:
    sheets = list_sheets(excel_bytes)   # fast: reads only the sheet names
except Exception as exc:
    st.error(f"Couldn't read that Excel file: {exc}")
    st.stop()


# ── STEP 1 — Excel layout (loads fast; only sheet names so far) ───────────────
st.subheader("1 · Excel layout")
st.caption(
    "Tell the tool where things live in the data sheet, then press Continue. "
    "Only the chosen sheet is scanned, which keeps things fast."
)

lc1, lc2, lc3, lc4 = st.columns([2, 1.4, 1, 1])
with lc1:
    sheet = st.selectbox(
        "Data sheet", sheets,
        index=sheets.index(engine.EXCEL_SHEET) if engine.EXCEL_SHEET in sheets else 0,
    )
with lc2:
    header_rows_raw = st.text_input(
        "Header rows", value=", ".join(str(r) for r in engine.HEADER_ROWS),
        help="Comma-separated rows holding the column hierarchy, e.g. 1, 2, 3, 4, 6",
    )
with lc3:
    letter_row_raw = st.text_input("Letter row", value=str(engine.LETTER_ROW),
                                   help="Row number holding the A/B/C letter codes.")
with lc4:
    label_col_raw = st.text_input("Label col", value=str(engine.LABEL_COL),
                                  help="Column number holding section/row labels.")

def _safe_int(raw, default):
    try:
        return int(str(raw).strip())
    except (ValueError, TypeError):
        return default

letter_row = _safe_int(letter_row_raw, engine.LETTER_ROW)
label_col  = _safe_int(label_col_raw, engine.LABEL_COL)

with st.expander("What the layout means"):
    st.code(
        "Row 1-4  : Hierarchical column headers (Country/Market -> Region -> Group -> Sub-col)\n"
        "Row 5    : 'Wave' label row\n"
        "Row 6    : 'Wave 4' / 'Wave 5' labels\n"
        "Row 7    : Letter codes (A, B, C, D ...)\n"
        "Row 8+   : Data, sig row immediately below each value row\n"
        "Column 1 : Section headers & row labels",
        language=None,
    )

header_rows = parse_int_list(header_rows_raw)
if not header_rows:
    st.error("Header rows must list at least one row number, e.g. 1, 2, 3, 4, 6.")
    st.stop()

# Staging: only scan the sheet after the user presses Continue (item 1).
# If any layout input changes, invalidate the previous scan so it re-runs.
layout_sig = (sheet, tuple(header_rows), letter_row, label_col)
if st.session_state.get("layout_sig") != layout_sig:
    st.session_state.pop("layout_ready", None)

if not st.session_state.get("layout_ready"):
    if st.button("Continue → scan this sheet", type="primary"):
        st.session_state["layout_ready"] = True
        st.session_state["layout_sig"] = layout_sig
        st.rerun()
    st.info("Set the layout above, then press Continue to scan the selected sheet.")
    st.stop()

# Build the hierarchy tree from the chosen layout (heavy step — show top bar)
_top_bar(True)
try:
    with st.spinner("Loading the workbook and scanning the selected sheet…"):
        # Load the workbook ONCE here; reused for the tree and the run below.
        _xl_hash = hash(excel_bytes)
        ws_loaded = load_worksheet(_xl_hash, excel_bytes, sheet)
        tree = build_tree(_xl_hash, (sheet, _xl_hash), tuple(header_rows),
                          int(letter_row), int(label_col), ws_loaded)
except Exception as exc:
    _top_bar(False)
    st.error(f"Couldn't read the header hierarchy with these layout settings: {exc}")
    st.stop()
_top_bar(False)

paths_raw = tree["paths"]
paths  = [p for (p, _col) in paths_raw]   # bare path tuples for cascading
levels = tree["levels"]
if not paths:
    st.error(
        "No Wave 4 / Wave 5 columns were found with these layout settings. "
        "Check the header rows and sheet."
    )
    st.stop()

st.success(f"Found {len(paths)} column cuts across {levels} hierarchy levels.")


# ── STEP 2 — Core settings (cascading dropdowns, unlocked after layout) ───────
st.subheader("2 · Default column cut")
st.caption(
    "Pick the column the tool uses for every shape unless overridden. Each level "
    "unlocks the next."
)

BLANK = "—"
chosen_default = []
cols = st.columns(levels)
for lvl in range(levels):
    opts = cascade_options(paths, chosen_default)
    if not opts:
        # No further levels apply for this branch — show a disabled placeholder
        with cols[lvl]:
            st.selectbox(f"Level {lvl + 1}", ["—"], index=0,
                         key=f"def_lvl{lvl}_disabled", disabled=True)
        chosen_default.append("")
        continue
    display = [BLANK if o == "" else o for o in opts]
    key = f"def_lvl{lvl}"
    # If the persisted choice is no longer valid under the parent selections,
    # drop it so the box falls back to the first valid option (true cascade).
    if key in st.session_state and st.session_state[key] not in display:
        del st.session_state[key]
    with cols[lvl]:
        sel = st.selectbox(f"Level {lvl + 1}", display, key=key)
    chosen_default.append("" if sel == BLANK else sel)

st.caption("Selected cut → " + " > ".join(c if c else "—" for c in chosen_default))

cset1, cset2 = st.columns(2)
with cset1:
    min_diff = st.number_input("Minimum % point difference", min_value=0, max_value=100,
                               value=engine.MIN_DIFF_PCT)
with cset2:
    skip_slides_raw = st.text_input(
        "Skip slides (comma-separated)",
        value=", ".join(str(s) for s in engine.SKIP_SLIDES))


# ── STEP 3 — Shape overrides ──────────────────────────────────────────────────
st.subheader("3 · Shape overrides (optional)")
st.caption(
    "Shapes that read from a different column than the default. Each row picks a "
    "full column path. Add a row, then choose its levels."
)

# Build level-value option lists for the editable table. To keep the table
# usable we offer ALL values seen at each level; invalid combinations are
# validated on run.
level_values = []
for lvl in range(levels):
    seen, vals = set(), []
    for p in paths:
        if lvl < len(p):
            v = p[lvl]
            disp = "—" if v == "" else v
            if disp not in seen:
                seen.add(disp); vals.append(disp)
    level_values.append(vals)

if "overrides_df" not in st.session_state:
    init_rows = []
    for entry in engine.SHAPE_OVERRIDES:
        # Entries may be legacy (slide, name, region, sub) or (slide, name, [path])
        if len(entry) == 4:
            s, name, c1, c2 = entry
            wanted = [v for v in (c1, c2) if v]
        else:
            s, name, path_list = entry
            wanted = [v for v in path_list if v]
        row = {"Slide": s, "Shape name": name}
        # Best-effort: find the full path whose values contain all wanted values
        match = next((p for p in paths if all(w in p for w in wanted)), None)
        for lvl in range(levels):
            if match is not None:
                row[f"Sub column {lvl + 1}"] = "—" if match[lvl] == "" else match[lvl]
            else:
                row[f"Sub column {lvl + 1}"] = ""
        init_rows.append(row)
    st.session_state.overrides_df = pd.DataFrame(init_rows)

col_cfg = {
    "Slide": st.column_config.NumberColumn("Slide", min_value=1, step=1, width="small"),
    "Shape name": st.column_config.TextColumn("Shape name", width="large"),
}
for lvl in range(levels):
    col_cfg[f"Sub column {lvl + 1}"] = st.column_config.SelectboxColumn(
        f"Sub column {lvl + 1}", options=[""] + level_values[lvl], width="medium")

overrides_df = st.data_editor(
    st.session_state.overrides_df,
    num_rows="dynamic", use_container_width=True, key="overrides_editor",
    column_config=col_cfg,
)


# ── STEP 4 — Run ──────────────────────────────────────────────────────────────
st.subheader("4 · Run")

have_result = st.session_state.get("run_result") is not None
auto_run = st.session_state.pop("rerun_requested", False)

# Show the run button only when there isn't a loaded result; once a run is
# loaded it stays on screen (downloads won't wipe it) until "Run again".
if not have_result:
    clicked = st.button("Apply triangles", type="primary", disabled=pptx_file is None)
    if pptx_file is None and not auto_run:
        st.info("Upload the PowerPoint to enable the run button.")
    run = clicked or auto_run
else:
    run = False

if run and pptx_file is not None:
    # Resolve overrides into (slide, shape, [path levels])
    overrides = []
    ov_errors = []
    for i, row in overrides_df.iterrows():
        slide = row.get("Slide")
        name  = (row.get("Shape name") or "")
        if pd.isna(slide) and not str(name).strip():
            continue
        if pd.isna(slide) or not str(name).strip():
            ov_errors.append(f"Override row {i + 1}: slide and shape name are required.")
            continue
        path_vals = []
        for lvl in range(levels):
            v = row.get(f"Sub column {lvl + 1}", "")
            v = "" if (v is None or (isinstance(v, float) and pd.isna(v)) or v == "—") else str(v).strip()
            path_vals.append(v)
        if not any(path_vals):
            ov_errors.append(f"Override row {i + 1}: choose at least one column level.")
            continue
        overrides.append((int(slide), str(name).strip(), path_vals))

    if ov_errors:
        for e in ov_errors:
            st.error(e)
        st.stop()

    # Push settings into the engine
    engine.EXCEL_SHEET     = sheet
    engine.MIN_DIFF_PCT    = int(min_diff)
    engine.SKIP_SLIDES     = parse_int_list(skip_slides_raw)
    engine.HEADER_ROWS     = header_rows
    engine.LETTER_ROW      = int(letter_row)
    engine.LABEL_COL       = int(label_col)
    engine.DEFAULT_CUT     = [c for c in chosen_default]
    engine.SHAPE_OVERRIDES = overrides
    nonblank = [c for c in chosen_default if c]
    engine.DEFAULT_REGION     = nonblank[0] if nonblank else ""
    engine.DEFAULT_SUB_COLUMN = nonblank[-1] if nonblank else ""

    log_lines = []
    _top_bar(True)
    try:
        with st.spinner("Processing…"):
            result = engine.run_sync(
                excel_bytes, pptx_file.getvalue(),
                log=log_lines.append, worksheet=ws_loaded)
    except Exception as exc:
        _top_bar(False)
        st.error(f"Something went wrong: {exc}")
        with st.expander("Show log"):
            st.code("\n".join(log_lines) or "(no output)")
        st.stop()
    _top_bar(False)

    # Persist so downloads (which rerun the script) don't wipe the results.
    st.session_state["run_result"] = {
        "total":        result["total"],
        "report_rows":  result["report_rows"],
        "pptx_bytes":   result["pptx_bytes"],
        "report_bytes": result["report_bytes"],
        "log":          "\n".join(log_lines),
        "pptx_name":    pptx_file.name.rsplit(".", 1)[0],
    }
    st.rerun()

# ── Render the loaded run result (persists across download reruns) ────────────
res = st.session_state.get("run_result")
if res is not None:
    rows = res["report_rows"]
    n_ok    = sum(1 for r in rows if r["status"] == engine.STATUS_OK)
    n_nosig = sum(1 for r in rows if r["status"] == engine.STATUS_NO_SIG)
    n_warn  = sum(1 for r in rows if r["status"] not in (engine.STATUS_OK, engine.STATUS_NO_SIG))

    top = st.columns([3, 1])
    with top[0]:
        st.success(f"Done — {res['total']} triangles applied across {len(rows)} shapes.")
    with top[1]:
        if st.button("↻ Run again", help="Clear this result and run again"):
            st.session_state.pop("run_result", None)
            st.session_state["rerun_requested"] = True
            st.rerun()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Triangles", res["total"])
    m2.metric("Stat testing applied", n_ok)
    m3.metric("Not significant", n_nosig)
    m4.metric("Data not found", n_warn)

    base = res["pptx_name"]
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "⬇ Updated PowerPoint", data=res["pptx_bytes"],
            file_name=f"{base} - updated.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            type="primary", key="dl_pptx")
    with d2:
        st.download_button(
            "⬇ Progress report (.xlsx)", data=res["report_bytes"],
            file_name="sync_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_report")

    # Colour-coded per-shape table (item 6) — dark-mode theme: deep fills with
    # light text so rows stay readable on a dark background.
    def _row_style(r):
        status = r["Status"]
        if status == engine.STATUS_OK:
            bg, fg = "#1E3B2F", "#A7E8C0"   # deep green / mint text
        elif status == engine.STATUS_NO_SIG:
            bg, fg = "#2B2F36", "#C7CDD6"   # slate grey / light grey text
        else:
            bg, fg = "#4A1F1F", "#F2B8B5"   # deep red / soft red text
        return [f"background-color: {bg}; color: {fg}"] * len(r)

    table_rows = [{"Slide": r["slide"], "Shape": r["shape_name"],
                   "Status": r["status"], "Applied": r.get("items_applied", "")}
                  for r in rows]
    with st.expander("Per-shape results", expanded=True):
        styled = (
            pd.DataFrame(table_rows).style
            .apply(_row_style, axis=1)
            .set_table_styles([
                {"selector": "th",
                 "props": [("background-color", "#15181D"),
                           ("color", "#E6E9EF"),
                           ("font-weight", "bold")]},
            ])
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)

    with st.expander("Processing log"):
        st.code(res["log"])
